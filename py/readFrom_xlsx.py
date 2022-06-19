# -*- coding: utf-8 -*-
import cx_Oracle as cx      #导入模块
import os
import openpyxl


os.environ['NLS_LANG'] = 'SIMPLIFIED CHINESE_CHINA.UTF8'


def my_main(i):
    con = cx.connect('kd_sale_dx', 'z007', '10.20.30.9:1521/racdb')  #创建连接
    cursor = con.cursor()       #创建游标

    #########

    print('START!')
    #1.get appno and appamount
    myExcel = openpyxl.load_workbook('E:\\wcy\\tmp\\backup.xlsx') #获取表格文件
    mySheet = myExcel.get_sheet_by_name('Sheet1') #获取指定的sheet

    select_sql = '''
    select discountno, upamount, downamount, ratevalue from sale_cfg_discountsection where discountno in (select discountno from sale_cfg_discount_paycenter where   fundcode >= '000000' and fundcode <= '000050' and (paycenterid = '0206' or paycenterid = '0102')) and valuetype = '0'
    '''
    cursor.execute(select_sql)
    data = cursor.fetchall()        #获取一条数据

    j = 1

    for rowsi in range(2, 207181):
        a1_cell_discountno = (mySheet.cell(row=rowsi,column=2)).value#省份 
        a1_cell_up = (mySheet.cell(row=rowsi,column=3)).value#省份 
        a1_cell_down = (mySheet.cell(row=rowsi,column=4)).value#省份 
        a1_cell_ratevalue = (mySheet.cell(row=rowsi,column=6)).value#省份 

        for data_row in data:
            if a1_cell_discountno == data_row[0] and a1_cell_up == data_row[1] and a1_cell_down == data_row[2]:
                print(f'''discountno : {a1_cell_discountno}, up : {a1_cell_up}, down : {a1_cell_down}, value_bak : {a1_cell_ratevalue}, value_db : {data_row[3]}''')
                if a1_cell_ratevalue != data_row[3]:
                    print(f"find different : {a1_cell_discountno} : total : {i}")
                    i = i + 1
                    if a1_cell_ratevalue == data_row[3] * 10:
                        print(f"update it : total : {j}")
                        j = j + 1
                        # input()

                        #1.2 update
                        update_sql = '''
                        update sale_cfg_discountsection set ratevalue = :var_ratevalue where discountno = :var_discountno and upamount = :var_upamount and downamount = :var_downamount and valuetype = '0'
                        '''
                        cursor.execute(update_sql, var_ratevalue = a1_cell_ratevalue, var_discountno = a1_cell_discountno, var_upamount = a1_cell_up, var_downamount = a1_cell_down)
                        print(f'''update sale_cfg_discountsection set ratevalue = {a1_cell_ratevalue} where discountno = {a1_cell_discountno} and upamount = {a1_cell_up} and downamount = {a1_cell_down} and valuetype = '0' ''')
                        if (cursor.rowcount == 1):
                            print("right! input 1 to commit, else to return")
                            if input() == "1":
                                con.commit()
                            else:
                                return

                        if (cursor.rowcount != 1):
                            print("error! input 2 to return")
                            if input() == "2":
                                return

    # #2.sql
    # query_sql = '''
    # '''

    # for fundcode in data:
    #     cursor.execute(query_sql, var_fundcode = fundcode[0])  #执行sql语句
    #     record = cursor.fetchall()        #获取一条数据
    #     print(str(fundcode[0]) + ' : ' + str(len(record)))
    #     # print(" : ")
    #     # print(len(record))

    #     #3. export to excel
    #     rows = record

    #     #获取字段名
    #     title = [ i[0] for i in cursor.description ]

    #     #创建excel表
    #     wb = openpyxl.Workbook()
    #     ws = wb.active

    #     #插入字段名到第一行
    #     for c in range(len(title)):
    #         ws.cell(1,c+1,value = title[c])

    #     #写入查询数据
    #     for r in range(len(rows)):
    #         for c in range(len(rows[r])):
    #             if rows[r][c]: #值不为空时写入，空值不写入
    #                 ws.cell(r+2,c+1,value=str(rows[r][c])) #str()防止用科学计数法写入造成信息丢失

    #     #保存sql脚本
    #     ws1 = wb.create_sheet('sql')
    #     ws1.cell(1,1,value=query_sql)

    #     wb.save('2403_' + fundcode[0] + '.xlsx')
    #     wb.close()

    #####################
    cursor.close()  #关闭游标
    con.close()     #关闭数据库连接

my_main(1)

# for i in range(0, 3):
#     my_main(i)
    # input()

# my_test_main()


