# -*- coding: utf-8 -*-
import cx_Oracle as cx      #导入模块
import os
import openpyxl


os.environ['NLS_LANG'] = 'SIMPLIFIED CHINESE_CHINA.UTF8'


def my_main(i):
    con = cx.connect('kd_sale_dx', 'z007', '10.20.30.9:1521/racdb')  #创建连接
    cursor = con.cursor()       #创建游标

    #########

    print('START!')
    #1.get fundcode
    get_fundcode = '''
    select distinct fundcode from v_bal_fund where custno = '2403' order by fundcode
    '''

    cursor.execute(get_fundcode)  #执行sql语句
    data = cursor.fetchall()        #获取一条数据

    #2.sql
    query_sql = '''
    SELECT appsheetserialno,
       taserialno,
       operdate,
       opertime,
       transactiondate,
       transactiontime,
       businesscode,
       fundcode,
       TARGETFUNDCODE,
       fundname,
       paytype,
       CASE
           WHEN applicationamount IS NULL
                OR applicationamount =0 THEN '--'
           WHEN applicationamount IS NOT NULL THEN to_char(applicationamount, 'FM99999999999990.00')
       END AS applicationamount,
       CASE
           WHEN applicationvol IS NULL
                OR applicationvol =0 THEN '--'
           WHEN applicationvol IS NOT NULL THEN to_char(applicationvol, 'FM99999999999990.00')
       END AS applicationvol,
       channelid,
       depositacct,
       bankname,
       CASE
           WHEN nav IS NULL
                OR nav =0 THEN '--'
           WHEN nav IS NOT NULL THEN to_char(nav, 'FM99999999999990.00')
       END AS nav,
       charge,
       CASE
           WHEN transamount IS NULL
                OR transamount =0 THEN '--'
           WHEN transamount IS NOT NULL THEN to_char(transamount)
       END AS transamount,
       amounttype,
       connectaccount,
       transstatus,
       confirmedamount,
       confirmedvol,
       status,
       paystatus,
       BUSINESSKIND,
       PAYMENTDATE,
       repaymentstatus,
       chkstatus
FROM
  (SELECT distinct(b.appsheetserialno),
          b.taserialno,
          b.operdate,
          b.opertime,
          b.transactiondate,
          b.transactiontime,
          b.businesscode,
          b.fundcode,
          b.TARGETFUNDCODE,
          b.fundname,
          b.paytype,
          b.applicationamount,
          b.applicationvol,
          b.channelid,
          c.depositacct,
          c.bankname,
          d.nav,
          b.charge,
          CASE
              WHEN b.businesscode in ('20',
                                      '22',
                                      '39') THEN to_char(bb.confirmedvol, 'FM99999999999990.00')
              WHEN b.businesscode in ('29',
                                      '59',
                                      '60',
                                      '61') THEN to_char(0.00, 'FM99999999999990.00')
              ELSE to_char(bb.confirmedamount, 'FM99999999999990.00')
          END AS transamount,
          CASE
              WHEN b.businesscode in ('20',
                                      '22',
                                      '39') THEN 'M'
              WHEN b.businesscode in ('29',
                                      '59',
                                      '60',
                                      '61') THEN '0.00'
              ELSE 'S'
          END AS amounttype,
          CASE
              WHEN b.businesscode in ('20',
                                      '22')
                   AND b.applicationamount<>
                     (SELECT t.applicationvol
                      FROM kd_sale_dx.v_app_trans t
                      WHERE t.appsheetserialno=b.referralcity) THEN '5'
              WHEN b.businesskind ='6' THEN '1'
              WHEN b.businesscode='20'
                   AND length(b.referralcity)=24 THEN '1'
              WHEN b.businesscode in ('20',
                                      '22')
                   AND b.applicationamount<>m.moneyasset
                   AND m.moneyasset=0 THEN '2'
              WHEN b.businesscode in ('20',
                                      '22')
                   AND b.applicationamount<>m.moneyasset
                   AND m.moneyasset<>0 THEN '3'
              ELSE '4'
          END AS connectaccount,
          CASE
              WHEN b.status='05' THEN '00'
              WHEN b.status = '08'
                   AND (b.returncode = '    '
                        OR b.returncode = '0000') THEN '01'
              WHEN (b.returncode !='0000'
                    AND b.returncode !='   ') THEN '02'
              WHEN (b.status='02'
                    AND b.returncode = '    '
                    AND b.paystatus='01'
                    AND b.paytype='1'
                    AND b.businesscode in ('20',
                                           '22')) THEN '04'
              WHEN (b.status in ('02',
                                 '06')
                    AND b.paystatus in ('  ',
                                        '01',
                                        '02')
                    AND b.returncode='    ') THEN '03'
              WHEN (b.businesscode='20'
                    AND b.status in ('02',
                                     '06',
                                     '07')
                    AND b.paystatus in ('  ',
                                        '01',
                                        '02')) THEN '03'
              WHEN (b.status in ('02',
                                 '04')
                    AND b.returncode = '    '
                    AND b.paystatus='03') THEN '04'
              WHEN b.businesscode in ('20',
                                      '22')
                   AND b.paystatus='00'
                   AND m.status='00'
                   AND b.status='02'
                   AND b.returncode = '    '
                   AND EXISTS
                     (SELECT 1
                      FROM kd_sale_dx.sale_bank_log g
                      WHERE g.appsheetserialno=b.appsheetserialno) THEN '04'
              WHEN (b.businesscode in ('29',
                                       '59')
                    AND b.status = '04') THEN '04'
              WHEN b.businesscode in ('20',
                                      '22')
                   AND b.paystatus='00'
                   AND m.status='00'
                   AND b.status in ('00',
                                    '01',
                                    '02')
                   AND floor(to_number(sysdate-to_date(b.transactiondate, 'yyyy-mm-dd')))<0 THEN '05'
              WHEN b.businesscode in ('20',
                                      '22')
                   AND b.paystatus='00'
                   AND m.status='00'
                   AND b.status in ('00',
                                    '01',
                                    '02')
                   AND floor(to_number(sysdate-to_date(b.transactiondate, 'yyyy-mm-dd')))=0
                   AND floor(to_number(to_date(to_char(sysdate, 'hh24:mi:ss'), 'hh24:mi:ss')-to_date('15:00:00', 'hh24:mi:ss'))*24*60*60)<0 THEN '05'
              WHEN b.businesscode in ('20',
                                      '22')
                   AND b.paystatus='00'
                   AND m.status='00'
                   AND b.status='04' THEN '06'
              WHEN b.businesscode in ('20',
                                      '22')
                   AND b.paystatus='00'
                   AND m.status='00'
                   AND b.status in ('00',
                                    '01',
                                    '02')
                   AND floor(to_number(sysdate-to_date(b.transactiondate, 'yyyy-mm-dd')))=0
                   AND floor(to_number(to_date(to_char(sysdate, 'hh24:mi:ss'), 'hh24:mi:ss')-to_date('15:00:00', 'hh24:mi:ss'))*24*60*60)>0 THEN '08'
              ELSE '07'
          END AS transstatus,
          bb.confirmedamount,
          bb.confirmedvol,
          b.status,
          b.PAYSTATUS,
          b.BUSINESSKIND,
          n.PAYMENTDATE,
          n.status AS repaymentstatus,
          n.chkstatus
   FROM kd_sale_dx.v_app_transnew b
   LEFT JOIN
     (SELECT *
      FROM kd_sale_dx.v_ack_trans
      WHERE detailflag='0') bb ON b.APPSHEETSERIALNO=bb.appsheetserialno
   LEFT JOIN kd_sale_dx.sale_acct_moneyaccount c ON c.custno = b.custno
   AND c.moneyaccount = b.moneyaccount
   LEFT JOIN kd_sale_dx.v_cfg_fundnav d ON d.fundcode= b.fundcode
   AND d.bulletindate = b.transactioncfmdate
   LEFT JOIN kd_sale_dx.v_app_payment m ON m.appsheetserialno=b.appsheetserialno
   LEFT JOIN kd_sale_dx.v_app_backbalance n ON n.ORGINAPPSHEETSERIALNO = b.appsheetserialno
   WHERE b.custno = 2403
     AND b.cancelflag = 'F'
     AND NOT EXISTS
       (SELECT 1
        FROM kd_sale_dx.sale_app_buycombination e
        WHERE e.combuyappsheetno = b.protocolno
          AND e.operdate > '20161226'
          AND e.custno = b.custno
          AND e.combinationcode BETWEEN 'ZH0004' AND 'ZHZ003')
     AND NOT EXISTS
       (SELECT 1
        FROM kd_sale_dx.sale_app_buycombination t
        WHERE t.combuyappsheetno = b.appsheetserialno
          AND t.custno = b.custno )
     AND b.businesscode in ('20',
                            '22',
                            '24',
                            '26',
                            '27',
                            '28',
                            '29',
                            '36',
                            '39',
                            '59',
                            '60',
                            '61')
     AND b.transactiondate<>'19900101'
   UNION SELECT distinct(a.appsheetserialno),
                a.taserialno,
                a.transactiondate AS operdate,
                a.transactiontime AS opertime,
                a.transactiondate,
                a.transactiontime,
                a.businesscode,
                a.fundcode,
                a.TARGETFUNDCODE,
                a.fundname,
                ' ' AS paytype,
                a.applicationamount,
                a.applicationvol,
                a.channelid,
                d.depositacct,
                d.bankname,
                a.nav,
                b.charge,
                CASE
                    WHEN a.businesscode='42' THEN to_char(a.applicationvol, 'FM99999999999990.00')
                    WHEN a.businesscode='43'
                         AND a.confirmedamount<>0 THEN to_char(a.confirmedamount, 'FM99999999999990.00')
                    WHEN a.businesscode='43'
                         AND a.confirmedvol<>0 THEN to_char(a.confirmedvol, 'FM99999999999990.00')
                    WHEN a.businesscode='44' THEN to_char(a.confirmedvol, 'FM99999999999990.00')
                    WHEN a.businesscode='45' THEN to_char(a.confirmedvol, 'FM99999999999990.00')
                    ELSE to_char(a.confirmedvol, 'FM99999999999990.00')
                END AS transamount,
                CASE
                    WHEN a.businesscode='42' THEN 'S'
                    WHEN a.businesscode='43'
                         AND a.confirmedvol<>0 THEN 'M'
                    WHEN a.businesscode='43'
                         AND a.confirmedamount<>0 THEN 'S'
                    WHEN a.businesscode='44' THEN 'S'
                    WHEN a.businesscode='45' THEN 'S'
                    ELSE 'S'
                END AS amounttype,
                '4' AS connectaccount,
                '01' AS transstatus,
                a.confirmedamount ,
                a.confirmedvol,
                '08' AS status,
                '  ' AS paystatus,
                '0' AS BUSINESSKIND,
                n.PAYMENTDATE,
                n.status AS repaymentstatus,
                n.chkstatus
   FROM kd_sale_dx.v_ack_transnew a
   LEFT JOIN kd_sale_dx.v_ack_trans b ON b.appsheetserialno=a.appsheetserialno
   AND b.detailflag='0'
   LEFT JOIN kd_sale_dx.v_app_backbalance n ON n.ORGINAPPSHEETSERIALNO = a.appsheetserialno
   LEFT JOIN kd_sale_dx.sale_acct_moneyaccount d ON d.custno = a.custno
   AND d.moneyaccount = a.moneyaccount
   WHERE a.custno = 2403
     AND a.businesscode in ('42',
                            '43',
                            '44',
                            '45')
     AND a.returncode='0000'
     AND a.detailflag='0' )
WHERE operdate IS NOT NULL
  AND operdate >= '19900101'
  AND operdate <= '20210424'
  AND fundcode = :var_fundcode
  OR targetfundcode= :var_fundcode
ORDER BY operdate DESC,
         opertime DESC
    '''

    for fundcode in data:
        cursor.execute(query_sql, var_fundcode = fundcode[0])  #执行sql语句
        record = cursor.fetchall()        #获取一条数据
        print(str(fundcode[0]) + ' : ' + str(len(record)))
        # print(" : ")
        # print(len(record))

        #3. export to excel
        rows = record

        #获取字段名
        title = [ i[0] for i in cursor.description ]

        #创建excel表
        wb = openpyxl.Workbook()
        ws = wb.active

        #插入字段名到第一行
        for c in range(len(title)):
            ws.cell(1,c+1,value = title[c])

        #写入查询数据
        for r in range(len(rows)):
            for c in range(len(rows[r])):
                if rows[r][c]: #值不为空时写入，空值不写入
                    ws.cell(r+2,c+1,value=str(rows[r][c])) #str()防止用科学计数法写入造成信息丢失

        #保存sql脚本
        ws1 = wb.create_sheet('sql')
        ws1.cell(1,1,value=query_sql)

        wb.save('2403_' + fundcode[0] + '.xlsx')
        wb.close()

    #####################
    cursor.close()  #关闭游标
    con.close()     #关闭数据库连接

my_main(1)

# for i in range(0, 3):
#     my_main(i)
    # input()

# my_test_main()


